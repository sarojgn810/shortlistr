"""
Excel Job Application Tracker — creates and updates Job_Application_Tracker.xlsx
Columns: Date Found | Source | Company | Title | Location | URL | Job ID |
         JD Snippet | Company Email | Status | Email Sent | Applied Date | Notes
"""

import os
import logging
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

from config import TRACKER_PATH

logger = logging.getLogger(__name__)

COLUMNS = [
    "Date Found", "Source", "Company", "Title", "Location",
    "URL", "Job ID", "Department", "JD Snippet",
    "Company Email", "Status", "Email Sent", "Applied Date", "Notes",
]

STATUS_COLORS = {
    "New":          "FFFDE7",   # light yellow
    "Applied":      "E8F5E9",   # light green
    "Interview":    "E3F2FD",   # light blue
    "Rejected":     "FFEBEE",   # light red
    "Offer":        "F3E5F5",   # light purple
    "Duplicate":    "F5F5F5",   # grey
}

HEADER_FILL = "1565C0"   # dark blue


def _col_widths():
    return {
        "Date Found": 12, "Source": 13, "Company": 22, "Title": 35,
        "Location": 20, "URL": 45, "Job ID": 20, "Department": 18,
        "JD Snippet": 60, "Company Email": 28, "Status": 12,
        "Email Sent": 12, "Applied Date": 14, "Notes": 35,
    }


def _create_workbook() -> "openpyxl.Workbook":
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"

    # Header row
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Column widths
    widths = _col_widths()
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 15)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # Stats sheet
    stats_ws = wb.create_sheet("Stats")
    stats_ws["A1"] = "Metric"
    stats_ws["B1"] = "Count"
    for cell in [stats_ws["A1"], stats_ws["B1"]]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)

    stats_data = [
        ("Total Applications", f"=COUNTA(Applications!A2:A10000)-COUNTIF(Applications!A2:A10000,\"\")"),
        ("Emails Sent",        f"=COUNTIF(Applications!L2:L10000,\"Yes\")"),
        ("Interviews",         f"=COUNTIF(Applications!K2:K10000,\"Interview\")"),
        ("Offers",             f"=COUNTIF(Applications!K2:K10000,\"Offer\")"),
        ("Rejected",           f"=COUNTIF(Applications!K2:K10000,\"Rejected\")"),
        ("New (Pending)",      f"=COUNTIF(Applications!K2:K10000,\"New\")"),
    ]
    for row_idx, (label, formula) in enumerate(stats_data, start=2):
        stats_ws[f"A{row_idx}"] = label
        stats_ws[f"B{row_idx}"] = formula
    stats_ws.column_dimensions["A"].width = 25
    stats_ws.column_dimensions["B"].width = 15

    return wb


def load_existing_ids() -> set:
    """Return set of job_ids already in the tracker (to deduplicate)."""
    if not os.path.exists(TRACKER_PATH) or not OPENPYXL_OK:
        return set()
    try:
        wb = openpyxl.load_workbook(TRACKER_PATH)
        ws = wb["Applications"]
        job_id_col = COLUMNS.index("Job ID") + 1
        ids = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[job_id_col - 1]:
                ids.add(str(row[job_id_col - 1]))
        return ids
    except Exception as e:
        logger.warning(f"Could not load existing tracker: {e}")
        return set()


def append_jobs(jobs: list[dict]) -> int:
    """Append new jobs to the tracker. Returns count of rows added."""
    if not OPENPYXL_OK:
        logger.error("openpyxl not installed — pip install openpyxl")
        return 0

    if not jobs:
        return 0

    # Load or create workbook
    if os.path.exists(TRACKER_PATH):
        wb = openpyxl.load_workbook(TRACKER_PATH)
        ws = wb["Applications"]
    else:
        wb = _create_workbook()
        ws = wb["Applications"]

    added = 0
    for job in jobs:
        row_data = [
            job.get("date_found", datetime.now().strftime("%Y-%m-%d")),
            job.get("source", ""),
            job.get("company", ""),
            job.get("title", ""),
            job.get("location", ""),
            job.get("url", ""),
            job.get("job_id", ""),
            job.get("department", ""),
            job.get("jd_snippet", "")[:500],
            job.get("company_email", ""),
            job.get("status", "New"),
            job.get("email_sent", "No"),
            job.get("applied_date", ""),
            job.get("notes", ""),
        ]
        row_num = ws.max_row + 1
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            # Row color by status
            status = job.get("status", "New")
            fill_color = STATUS_COLORS.get(status, "FFFFFF")
            cell.fill = PatternFill("solid", fgColor=fill_color)

        ws.row_dimensions[row_num].height = 40
        added += 1

    wb.save(TRACKER_PATH)
    logger.info(f"Tracker updated: +{added} rows → {TRACKER_PATH}")
    return added


def update_job_status(job_id: str, status: str, email_sent: str = None, notes: str = None):
    """Update status/email_sent for a specific job_id in the tracker."""
    if not OPENPYXL_OK or not os.path.exists(TRACKER_PATH):
        return
    try:
        wb = openpyxl.load_workbook(TRACKER_PATH)
        ws = wb["Applications"]
        job_id_col = COLUMNS.index("Job ID") + 1
        status_col  = COLUMNS.index("Status") + 1
        email_col   = COLUMNS.index("Email Sent") + 1
        applied_col = COLUMNS.index("Applied Date") + 1
        notes_col   = COLUMNS.index("Notes") + 1

        for row in ws.iter_rows(min_row=2):
            if str(row[job_id_col - 1].value) == str(job_id):
                row[status_col - 1].value = status
                if email_sent:
                    row[email_col - 1].value = email_sent
                    row[applied_col - 1].value = datetime.now().strftime("%Y-%m-%d")
                if notes:
                    row[notes_col - 1].value = notes
                # Update row color
                fill_color = STATUS_COLORS.get(status, "FFFFFF")
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=fill_color)
                break

        wb.save(TRACKER_PATH)
    except Exception as e:
        logger.error(f"Error updating tracker for job {job_id}: {e}")
